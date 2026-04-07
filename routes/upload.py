"""
Flexible multi-format bulk upload system.

Three endpoints under the upload blueprint:
  POST /admin/upload/candidates  — upsert into people_profiles
  POST /admin/upload/clients     — upsert into client_contacts
  GET  /admin/upload/stats       — pipeline-wide counts

All three are @require_admin. CSV and XLSX files are auto-detected
by extension; the caller may also supply a column_map JSON that
overrides the default alias mapping for non-standard file layouts.
"""
import json
import re
from datetime import datetime, timezone
from typing import Any, Optional

from flask import Blueprint, request

from utils.auth_helpers import require_admin
from utils.response_helpers import ok, bad
from config.clients import supabase_client


upload_bp = Blueprint("upload", __name__)


# ── Canonical header alias map ───────────────────────────────────────────────
# Keys are normalised headers (strip, lowercase, punct -> underscore).
# Values are our canonical field names. User-supplied column_map overrides
# this at call time.
_HEADER_ALIASES = {
    # Name
    "name": "name",
    "full_name": "name",
    "contact_name": "name",
    "first_name": "first_name",
    "firstname": "first_name",
    "given_name": "first_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "surname": "last_name",
    "family_name": "last_name",
    # Email
    "email": "email",
    "email_address": "email",
    "e_mail": "email",
    "e_mail_address": "email",
    # Phones
    "phone": "phone",
    "phone_number": "phone",
    "mobile": "mobile",
    "mobile_phone": "mobile",
    "mobile_number": "mobile",
    "cell": "mobile",
    "cell_phone": "mobile",
    "work_phone": "work_phone",
    "office_phone": "work_phone",
    "business_phone": "work_phone",
    "home_phone": "home_phone",
    # Title / role
    "title": "title",
    "job_title": "title",
    "current_job_title": "title",
    "position": "title",
    "role": "title",
    "headline": "title",
    # Company
    "company": "company",
    "company_name": "company",
    "organisation": "company",
    "organization": "company",
    "employer": "company",
    # Location
    "location": "location",
    "desired_locations": "location",
    "city": "location",
    "country": "location",
    # Misc useful extras
    "notes": "notes",
    "comments": "notes",
    "linkedin_url": "linkedin_url",
    "linkedin": "linkedin_url",
    "connected_on": "connected_on",
}


def _normalise_header(raw: Any) -> str:
    if not isinstance(raw, str):
        return ""
    return re.sub(r"[^a-z0-9]+", "_", raw.strip().lower()).strip("_")


def _build_row_mapper(
    headers: list[str],
    user_column_map: Optional[dict] = None,
) -> dict:
    """
    Return a dict {canonical_field: header_index} by combining the
    default alias map with the user's column_map override.

    user_column_map has the shape:
        {"canonical_field": "Exact Header From File"}
    e.g. {"email": "E-Mail Address", "mobile": "Mob."}
    """
    mapping: dict = {}

    # Pass 1 — default aliases
    for idx, header in enumerate(headers):
        norm = _normalise_header(header)
        canonical = _HEADER_ALIASES.get(norm)
        if canonical and canonical not in mapping:
            mapping[canonical] = idx

    # Pass 2 — user overrides win
    if user_column_map and isinstance(user_column_map, dict):
        # Build a case-insensitive index of the raw headers for exact lookup
        header_lookup = {(h or "").strip().lower(): i for i, h in enumerate(headers)}
        for canonical_field, raw_target in user_column_map.items():
            if not isinstance(canonical_field, str) or not isinstance(raw_target, str):
                continue
            idx = header_lookup.get(raw_target.strip().lower())
            if idx is not None:
                mapping[canonical_field] = idx
    return mapping


def _extract_row(row_values: list, mapping: dict) -> dict:
    """Apply the mapping to a raw row; returns a dict of canonical values."""
    out: dict = {}
    for canonical, idx in mapping.items():
        if idx < len(row_values):
            v = row_values[idx]
            if v is None:
                continue
            if isinstance(v, str):
                v = v.strip()
                if v:
                    out[canonical] = v
            else:
                out[canonical] = v
    return out


def _split_full_name(full: str) -> tuple:
    if not isinstance(full, str):
        return (None, None)
    parts = full.strip().split(None, 1)
    if not parts:
        return (None, None)
    if len(parts) == 1:
        return (parts[0], None)
    return (parts[0], parts[1])


def _normalise_phone_multi(raw: Any) -> Optional[str]:
    """
    Smart phone normalisation with Irish (+353) and UK (+44) heuristics.

    Rules:
      - Strip spaces, dashes, parens, dots
      - Already +N{8..15} -> return as-is
      - Starts with "00" -> swap for "+"
      - Irish mobile: starts with "08" -> "+353" + remainder after leading 0
      - UK mobile: starts with "07" -> "+44" + remainder after leading 0
      - Otherwise: prepend "+" and validate
      - Invalid (wrong length, non-digit) -> None
    """
    if not isinstance(raw, str):
        raw = str(raw) if raw is not None else ""
    cleaned = re.sub(r"[\s\-().]", "", raw.strip())
    if not cleaned:
        return None

    if cleaned.startswith("+"):
        if re.fullmatch(r"\+\d{8,15}", cleaned):
            return cleaned
        return None

    if cleaned.startswith("00"):
        cleaned = "+" + cleaned[2:]
        return cleaned if re.fullmatch(r"\+\d{8,15}", cleaned) else None

    # Irish mobile — 08X numbers
    if re.fullmatch(r"08\d{7,9}", cleaned):
        return "+353" + cleaned[1:]

    # UK mobile — 07 numbers (10-11 digits total)
    if re.fullmatch(r"07\d{9}", cleaned):
        return "+44" + cleaned[1:]

    # Bare digits — best effort
    if re.fullmatch(r"\d{8,15}", cleaned):
        return "+" + cleaned

    return None


# ── File reading (CSV + XLSX) ────────────────────────────────────────────────

def _read_rows_from_upload(file_field) -> tuple:
    """
    Return (headers, rows) where headers is a list[str] and rows is
    a list[list[Any]] (raw values in header order). Auto-detects format
    by filename extension. Raises ValueError with a human-readable
    message on failure.
    """
    filename = (file_field.filename or "").lower()
    raw_bytes = file_field.read()

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
            import io as _io
            wb = load_workbook(
                filename=_io.BytesIO(raw_bytes),
                read_only=True,
                data_only=True,
            )
            ws = wb.active
            if ws is None:
                raise ValueError("Empty workbook")
            iterator = ws.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration:
                return ([], [])
            headers = [str(h) if h is not None else "" for h in header_row]
            rows = []
            for row in iterator:
                rows.append(list(row))
            return (headers, rows)
        except ImportError:
            raise ValueError("openpyxl is not installed — XLSX files cannot be read")
        except Exception as e:
            raise ValueError(f"Could not parse XLSX: {e}")

    if filename.endswith(".xls"):
        raise ValueError(".xls (legacy Excel) is not supported — please save as .xlsx or .csv")

    # Default: CSV
    try:
        import csv
        import io as _io
        text = raw_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(_io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            return ([], [])
        headers = [(h or "").strip() for h in all_rows[0]]
        rows = all_rows[1:]
        return (headers, rows)
    except Exception as e:
        raise ValueError(f"Could not parse CSV: {e}")


# ── Supabase helpers ─────────────────────────────────────────────────────────

def _upsert_channel_identity_best_effort(
    user_id: Optional[str],
    channel: str,
    value: str,
    row_index: int,
) -> None:
    """Only inserts when we have a user_id (schema requires non-null)."""
    if not user_id or not supabase_client or not value:
        return
    try:
        supabase_client.table("channel_identities").insert({
            "user_id": user_id,
            "channel": channel,
            "value": value,
        }).execute()
    except Exception as e:
        print(f"[UPLOAD] channel_identities {channel} insert failed row={row_index}: {e}", flush=True)


# ── Endpoint: POST /admin/upload/candidates ──────────────────────────────────

@upload_bp.route("/admin/upload/candidates", methods=["POST"])
@require_admin
def upload_candidates():
    """
    POST /admin/upload/candidates
    multipart/form-data:
      file:       CSV or XLSX
      source:     one of linkedin_connections, candidate_database,
                  hr_contacts, manual (free-form; stored verbatim)
      column_map: optional JSON string mapping canonical field names
                  to the exact header strings used in the file

    Returns {inserted, updated, skipped, errors[:50], has_phones,
             has_emails, total_processed, source}.
    """
    if not supabase_client:
        return bad("Database not available", 503)

    file_field = request.files.get("file")
    if file_field is None or not file_field.filename:
        return bad("file field is required", 400)

    source = (request.form.get("source") or "manual").strip() or "manual"
    user_column_map = _parse_column_map_arg(request.form.get("column_map"))

    try:
        headers, rows = _read_rows_from_upload(file_field)
    except ValueError as e:
        return bad(str(e), 400)

    if not headers or not rows:
        return bad("Empty file or no data rows", 400)

    mapping = _build_row_mapper(headers, user_column_map)
    if not mapping:
        return bad(
            "Could not map any known columns from the file. "
            "Supply column_map to override, or check the file headers.",
            400,
        )

    print(
        f"[UPLOAD] candidates source={source} filename={file_field.filename!r} "
        f"headers={headers} mapping={mapping} rows={len(rows)}",
        flush=True,
    )

    now_iso = datetime.now(timezone.utc).isoformat()
    inserted = 0
    updated = 0
    skipped = 0
    has_phones = 0
    has_emails = 0
    errors: list = []
    processed = 0

    for row_index, raw_row in enumerate(rows, start=2):  # row 1 = header
        processed += 1
        try:
            data = _extract_row(raw_row, mapping)

            first_name = data.get("first_name")
            last_name = data.get("last_name")
            if not first_name and not last_name and data.get("name"):
                first_name, last_name = _split_full_name(data["name"])

            email = data.get("email")
            if isinstance(email, str):
                email = email.strip().lower() or None

            mobile_raw = data.get("mobile") or data.get("phone")
            work_phone_raw = data.get("work_phone")
            mobile = _normalise_phone_multi(mobile_raw) if mobile_raw else None
            work_phone = _normalise_phone_multi(work_phone_raw) if work_phone_raw else None

            title = data.get("title")
            company = data.get("company")
            location = data.get("location")
            linkedin_url = data.get("linkedin_url")

            if not any([first_name, last_name, email, linkedin_url]):
                skipped += 1
                errors.append({"row": row_index, "reason": "no_identifying_fields"})
                continue

            headline = None
            if title and company:
                headline = f"{title} at {company}"
            elif title:
                headline = title
            elif company:
                headline = company

            sm = {
                "original_company": company,
                "upload_date": now_iso,
                "upload_source": source,
                "upload_filename": file_field.filename,
                "has_mobile": bool(mobile),
                "has_work_phone": bool(work_phone),
            }
            if email:
                sm["upload_email"] = email
            if mobile:
                sm["upload_phone"] = mobile
            if work_phone:
                sm["upload_work_phone"] = work_phone
            if data.get("connected_on"):
                sm["linkedin_connected_on"] = data["connected_on"]
            if data.get("notes"):
                sm["upload_notes"] = data["notes"]

            existing_row = None
            if email:
                try:
                    dup_resp = (
                        supabase_client.table("people_profiles")
                        .select("id, first_name, last_name, headline, location, "
                                "linkedin_url, source_metadata, user_id, approved")
                        .eq("source_metadata->>upload_email", email)
                        .limit(1)
                        .execute()
                    )
                    if dup_resp.data:
                        existing_row = dup_resp.data[0]
                except Exception as e:
                    print(f"[UPLOAD] email dedup failed row={row_index}: {e}", flush=True)

            if not existing_row and linkedin_url:
                try:
                    dup_resp = (
                        supabase_client.table("people_profiles")
                        .select("id, first_name, last_name, headline, location, "
                                "linkedin_url, source_metadata, user_id, approved")
                        .eq("linkedin_url", linkedin_url)
                        .limit(1)
                        .execute()
                    )
                    if dup_resp.data:
                        existing_row = dup_resp.data[0]
                except Exception as e:
                    print(f"[UPLOAD] linkedin dedup failed row={row_index}: {e}", flush=True)

            if existing_row:
                if existing_row.get("approved") is True:
                    skipped += 1
                    errors.append({"row": row_index, "reason": "already_approved"})
                    continue

                merged_sm = dict(existing_row.get("source_metadata") or {})
                merged_sm.update(sm)

                update_payload: dict = {"source_metadata": merged_sm}
                if first_name and not existing_row.get("first_name"):
                    update_payload["first_name"] = first_name
                if last_name and not existing_row.get("last_name"):
                    update_payload["last_name"] = last_name
                if headline and not existing_row.get("headline"):
                    update_payload["headline"] = headline
                if location and not existing_row.get("location"):
                    update_payload["location"] = location
                if linkedin_url and not existing_row.get("linkedin_url"):
                    update_payload["linkedin_url"] = linkedin_url

                try:
                    supabase_client.table("people_profiles").update(update_payload).eq(
                        "id", existing_row["id"]
                    ).execute()
                    updated += 1
                except Exception as e:
                    skipped += 1
                    errors.append({"row": row_index, "reason": f"update_failed: {e}"})
                    continue

                _upsert_channel_identity_best_effort(
                    existing_row.get("user_id"), "email", email, row_index,
                ) if email else None
                if mobile:
                    _upsert_channel_identity_best_effort(
                        existing_row.get("user_id"), "phone", mobile, row_index,
                    )
                elif work_phone:
                    _upsert_channel_identity_best_effort(
                        existing_row.get("user_id"), "phone", work_phone, row_index,
                    )
            else:
                insert_payload = {
                    "first_name": first_name,
                    "last_name": last_name,
                    "headline": headline,
                    "location": location,
                    "linkedin_url": linkedin_url,
                    "approved": False,
                    "source": source,
                    "source_metadata": sm,
                }
                try:
                    resp = supabase_client.table("people_profiles").insert(insert_payload).execute()
                    if resp.data:
                        inserted += 1
                    else:
                        skipped += 1
                        errors.append({"row": row_index, "reason": "insert_returned_no_data"})
                        continue
                except Exception as e:
                    skipped += 1
                    errors.append({"row": row_index, "reason": f"insert_failed: {e}"})
                    continue

            if email:
                has_emails += 1
            if mobile or work_phone:
                has_phones += 1

            if processed % 100 == 0:
                print(
                    f"[UPLOAD] processed={processed}/{len(rows)} "
                    f"inserted={inserted} updated={updated} skipped={skipped}",
                    flush=True,
                )
        except Exception as e:
            skipped += 1
            errors.append({"row": row_index, "reason": f"row_error: {e}"})

    print(
        f"[UPLOAD] DONE candidates source={source} total={processed} "
        f"inserted={inserted} updated={updated} skipped={skipped} "
        f"has_phones={has_phones} has_emails={has_emails}",
        flush=True,
    )

    return ok({
        "source": source,
        "total_processed": processed,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "has_phones": has_phones,
        "has_emails": has_emails,
        "errors": errors[:50],
        "errors_truncated": len(errors) > 50,
    }, status=200)


# ── Endpoint: POST /admin/upload/clients ─────────────────────────────────────

@upload_bp.route("/admin/upload/clients", methods=["POST"])
@require_admin
def upload_clients():
    """
    POST /admin/upload/clients
    multipart/form-data:
      file:       CSV or XLSX
      source:     free-form tag stored on the row
      column_map: optional JSON override

    Upserts by email into client_contacts. Returns the same shape
    as /admin/upload/candidates.
    """
    if not supabase_client:
        return bad("Database not available", 503)

    file_field = request.files.get("file")
    if file_field is None or not file_field.filename:
        return bad("file field is required", 400)

    source = (request.form.get("source") or "manual").strip() or "manual"
    user_column_map = _parse_column_map_arg(request.form.get("column_map"))

    try:
        headers, rows = _read_rows_from_upload(file_field)
    except ValueError as e:
        return bad(str(e), 400)

    if not headers or not rows:
        return bad("Empty file or no data rows", 400)

    mapping = _build_row_mapper(headers, user_column_map)
    if not mapping:
        return bad("Could not map any known columns. Supply column_map to override.", 400)

    print(
        f"[UPLOAD] clients source={source} filename={file_field.filename!r} "
        f"headers={headers} mapping={mapping} rows={len(rows)}",
        flush=True,
    )

    now_iso = datetime.now(timezone.utc).isoformat()
    inserted = 0
    updated = 0
    skipped = 0
    has_phones = 0
    has_emails = 0
    errors: list = []
    processed = 0

    for row_index, raw_row in enumerate(rows, start=2):
        processed += 1
        try:
            data = _extract_row(raw_row, mapping)

            first_name = data.get("first_name")
            last_name = data.get("last_name")
            name = data.get("name")
            if not name and (first_name or last_name):
                name = " ".join(p for p in (first_name, last_name) if p).strip()
            elif name and not first_name and not last_name:
                first_name, last_name = _split_full_name(name)

            email = data.get("email")
            if isinstance(email, str):
                email = email.strip().lower() or None

            mobile = _normalise_phone_multi(data.get("mobile") or data.get("phone") or "")
            work_phone = _normalise_phone_multi(data.get("work_phone") or "")
            title = data.get("title")
            company = data.get("company")
            notes = data.get("notes")

            if not any([name, email, company]):
                skipped += 1
                errors.append({"row": row_index, "reason": "no_identifying_fields"})
                continue

            row_payload = {
                "name": name,
                "title": title,
                "company": company,
                "email": email,
                "work_phone": work_phone,
                "mobile": mobile,
                "source": source,
                "notes": notes,
                "source_metadata": {
                    "upload_date": now_iso,
                    "upload_filename": file_field.filename,
                    "has_mobile": bool(mobile),
                    "has_work_phone": bool(work_phone),
                },
            }

            existing = None
            if email:
                try:
                    dup_resp = (
                        supabase_client.table("client_contacts")
                        .select("id")
                        .eq("email", email)
                        .limit(1)
                        .execute()
                    )
                    if dup_resp.data:
                        existing = dup_resp.data[0]
                except Exception as e:
                    print(f"[UPLOAD] client email dedup failed row={row_index}: {e}", flush=True)

            if existing:
                try:
                    update_payload = dict(row_payload)
                    update_payload.pop("source_metadata", None)
                    update_payload["updated_at"] = now_iso
                    # Don't overwrite name/company/title with None if they were
                    # populated previously — fill-empty merge on the key text
                    # fields only.
                    update_payload = {k: v for k, v in update_payload.items() if v is not None}
                    supabase_client.table("client_contacts").update(update_payload).eq(
                        "id", existing["id"]
                    ).execute()
                    updated += 1
                except Exception as e:
                    skipped += 1
                    errors.append({"row": row_index, "reason": f"update_failed: {e}"})
                    continue
            else:
                try:
                    resp = supabase_client.table("client_contacts").insert(row_payload).execute()
                    if resp.data:
                        inserted += 1
                    else:
                        skipped += 1
                        errors.append({"row": row_index, "reason": "insert_returned_no_data"})
                        continue
                except Exception as e:
                    skipped += 1
                    errors.append({"row": row_index, "reason": f"insert_failed: {e}"})
                    continue

            if email:
                has_emails += 1
            if mobile or work_phone:
                has_phones += 1

            if processed % 100 == 0:
                print(
                    f"[UPLOAD] clients processed={processed}/{len(rows)} "
                    f"inserted={inserted} updated={updated} skipped={skipped}",
                    flush=True,
                )
        except Exception as e:
            skipped += 1
            errors.append({"row": row_index, "reason": f"row_error: {e}"})

    print(
        f"[UPLOAD] DONE clients source={source} total={processed} "
        f"inserted={inserted} updated={updated} skipped={skipped} "
        f"has_phones={has_phones} has_emails={has_emails}",
        flush=True,
    )

    return ok({
        "source": source,
        "total_processed": processed,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "has_phones": has_phones,
        "has_emails": has_emails,
        "errors": errors[:50],
        "errors_truncated": len(errors) > 50,
    }, status=200)


# ── Endpoint: GET /admin/upload/stats ────────────────────────────────────────

@upload_bp.route("/admin/upload/stats", methods=["GET"])
@require_admin
def upload_stats():
    """
    GET /admin/upload/stats

    Pipeline-wide counts for the admin dashboard.
    """
    if not supabase_client:
        return bad("Database not available", 503)

    candidates_report = {
        "total": 0,
        "by_source": {},
        "with_phone": 0,
        "with_email": 0,
        "approved": 0,
        "called": 0,
    }
    clients_report = {
        "total": 0,
        "with_phone": 0,
        "with_email": 0,
        "by_outreach_status": {},
    }

    # ── Candidates (people_profiles) ────────────────────────────────────
    try:
        total_resp = (
            supabase_client.table("people_profiles")
            .select("id", count="exact")
            .limit(1)
            .execute()
        )
        candidates_report["total"] = total_resp.count or 0
    except Exception as e:
        print(f"[UPLOAD-STATS] candidates total failed: {e}", flush=True)

    # by_source — fetch distinct source values via a scan. supabase-py
    # doesn't expose GROUP BY so we paginate a lightweight select.
    try:
        resp = (
            supabase_client.table("people_profiles")
            .select("source, source_metadata, approved, user_id")
            .limit(5000)
            .execute()
        )
        rows = resp.data or []
        by_source: dict = {}
        with_email = 0
        with_phone = 0
        approved_count = 0
        for r in rows:
            src = r.get("source") or "unknown"
            by_source[src] = by_source.get(src, 0) + 1
            sm = r.get("source_metadata") or {}
            if sm.get("upload_email") or sm.get("enriched_email"):
                with_email += 1
            if sm.get("upload_phone") or sm.get("enriched_phone"):
                with_phone += 1
            if r.get("approved") is True:
                approved_count += 1
        candidates_report["by_source"] = by_source
        candidates_report["with_email"] = with_email
        candidates_report["with_phone"] = with_phone
        candidates_report["approved"] = approved_count
    except Exception as e:
        print(f"[UPLOAD-STATS] candidates scan failed: {e}", flush=True)

    # called — count completed outbound_call_jobs
    try:
        called_resp = (
            supabase_client.table("outbound_call_jobs")
            .select("id", count="exact")
            .eq("status", "completed")
            .limit(1)
            .execute()
        )
        candidates_report["called"] = called_resp.count or 0
    except Exception as e:
        print(f"[UPLOAD-STATS] candidates called count failed: {e}", flush=True)

    # ── Clients (client_contacts) ───────────────────────────────────────
    try:
        total_resp = (
            supabase_client.table("client_contacts")
            .select("id", count="exact")
            .limit(1)
            .execute()
        )
        clients_report["total"] = total_resp.count or 0
    except Exception as e:
        print(f"[UPLOAD-STATS] clients total failed (table may not exist yet): {e}", flush=True)

    try:
        resp = (
            supabase_client.table("client_contacts")
            .select("email, work_phone, mobile, outreach_status")
            .limit(5000)
            .execute()
        )
        rows = resp.data or []
        with_email = 0
        with_phone = 0
        by_status: dict = {}
        for r in rows:
            if r.get("email"):
                with_email += 1
            if r.get("mobile") or r.get("work_phone"):
                with_phone += 1
            status = r.get("outreach_status") or "not_contacted"
            by_status[status] = by_status.get(status, 0) + 1
        clients_report["with_email"] = with_email
        clients_report["with_phone"] = with_phone
        clients_report["by_outreach_status"] = by_status
    except Exception as e:
        print(f"[UPLOAD-STATS] clients scan failed (table may not exist yet): {e}", flush=True)

    return ok({
        "candidates": candidates_report,
        "clients": clients_report,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }, status=200)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_column_map_arg(raw: Optional[str]) -> Optional[dict]:
    """Parse the column_map form field — tolerant of None/empty/bad JSON."""
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else None
    except Exception as e:
        print(f"[UPLOAD] column_map JSON parse failed: {e}", flush=True)
        return None
